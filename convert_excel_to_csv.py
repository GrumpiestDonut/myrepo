import csv
import os
from pyxlsb import open_workbook
from openpyxl import load_workbook
import xlrd
from tqdm import tqdm

def convert_xlsb_to_csv(input_path, output_path):
    with open_workbook(input_path) as workbook:
        sheet_name = workbook.sheets[0]
        sheet = workbook.get_sheet(sheet_name)
        
        with open(output_path, mode='w', newline='') as file:
            writer = csv.writer(file)
            total_rows = sum(1 for _ in sheet.rows())
            for row in tqdm(sheet.rows(), total=total_rows, desc=f"Converting {os.path.basename(input_path)}"):
                writer.writerow([item.v for item in row])
    
    print(f"Conversion complete for {input_path}!")

def convert_xlsx_to_csv(input_path, output_path):
    workbook = load_workbook(input_path, read_only=True)
    sheet = workbook.active
    
    with open(output_path, mode='w', newline='') as file:
        writer = csv.writer(file)
        total_rows = sheet.max_row
        for row in tqdm(sheet.iter_rows(values_only=True), total=total_rows, desc=f"Converting {os.path.basename(input_path)}"):
            writer.writerow(row)
    
    print(f"Conversion complete for {input_path}!")

def convert_xls_to_csv(input_path, output_path):
    workbook = xlrd.open_workbook(input_path)
    sheet = workbook.sheet_by_index(0)
    
    with open(output_path, mode='w', newline='') as file:
        writer = csv.writer(file)
        total_rows = sheet.nrows
        for row_idx in tqdm(range(total_rows), total=total_rows, desc=f"Converting {os.path.basename(input_path)}"):
            writer.writerow(sheet.row_values(row_idx))
    
    print(f"Conversion complete for {input_path}!")

def main():
    folder_path = r"PATHINPUT"
    output_folder = r"PATHOUTPUT"
    
    for filename in os.listdir(folder_path):
        input_path = os.path.join(folder_path, filename)
        output_path = os.path.join(output_folder, f"{os.path.splitext(filename)[0]}.csv")
        
        if filename.endswith(".xlsb"):
            convert_xlsb_to_csv(input_path, output_path)
        elif filename.endswith(".xlsx"):
            convert_xlsx_to_csv(input_path, output_path)
        elif filename.endswith(".xls"):
            convert_xls_to_csv(input_path, output_path)

if __name__ == "__main__":
    main()
